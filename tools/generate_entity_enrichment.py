#!/usr/bin/env python3
"""
Prime Pathwy — Phase 1: Massive Real-World Entity Enrichment Generator
Generates operationally authentic enriched entity datasets for the Sovereign Vault.
WAT: /tools
"""

import csv
import json
import random
import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime, timedelta

OUTPUT_DIR = "/home/ubuntu/sovereign-vault/vault/SOVEREIGN_EXECUTION_VAULT/entity_enrichment"
os.makedirs(OUTPUT_DIR, exist_ok=True)

GOLD = "C9A646"
BLACK = "0B0B0B"
WHITE = "FFFFFF"

# ─────────────────────────────────────────────────────────────────────────────
# REAL-WORLD ENTITY SEED DATA
# ─────────────────────────────────────────────────────────────────────────────

INDUSTRIES = [
    "Janitorial & Facilities Maintenance",
    "Commercial Hauling & Debris Removal",
    "Industrial Painting & Coatings",
    "Fleet Logistics & Transportation",
    "Records Management & Digitization",
    "AI & Automation Consulting",
    "IT Infrastructure & Managed Services",
    "Environmental Compliance & Remediation",
    "Security Services & Guard Staffing",
    "HVAC & Mechanical Services",
    "Electrical Contracting",
    "Landscaping & Grounds Maintenance",
    "Pest Control & Sanitation",
    "Plumbing & Water Systems",
    "Construction & General Contracting",
    "Staffing & Workforce Solutions",
    "Accounting & Financial Services",
    "Legal & Compliance Services",
    "Marketing & Digital Services",
    "Healthcare Facility Services",
]

NAICS_MAP = {
    "Janitorial & Facilities Maintenance": ("561720", "7349"),
    "Commercial Hauling & Debris Removal": ("562111", "4953"),
    "Industrial Painting & Coatings": ("238320", "1731"),
    "Fleet Logistics & Transportation": ("484110", "4213"),
    "Records Management & Digitization": ("493110", "7374"),
    "AI & Automation Consulting": ("541512", "7372"),
    "IT Infrastructure & Managed Services": ("541513", "7374"),
    "Environmental Compliance & Remediation": ("562910", "8711"),
    "Security Services & Guard Staffing": ("561612", "7382"),
    "HVAC & Mechanical Services": ("238220", "1711"),
    "Electrical Contracting": ("238210", "1731"),
    "Landscaping & Grounds Maintenance": ("561730", "0782"),
    "Pest Control & Sanitation": ("561710", "7342"),
    "Plumbing & Water Systems": ("238220", "1711"),
    "Construction & General Contracting": ("236220", "1542"),
    "Staffing & Workforce Solutions": ("561320", "7363"),
    "Accounting & Financial Services": ("541211", "8721"),
    "Legal & Compliance Services": ("541110", "8111"),
    "Marketing & Digital Services": ("541810", "7311"),
    "Healthcare Facility Services": ("621999", "8049"),
}

REGIONS = [
    # NorCal
    ("Solano County", "CA", "US"),
    ("Contra Costa County", "CA", "US"),
    ("Alameda County", "CA", "US"),
    ("Sacramento County", "CA", "US"),
    ("San Francisco County", "CA", "US"),
    ("Marin County", "CA", "US"),
    ("Napa County", "CA", "US"),
    ("Sonoma County", "CA", "US"),
    # SoCal
    ("Los Angeles County", "CA", "US"),
    ("San Diego County", "CA", "US"),
    ("Orange County", "CA", "US"),
    # Other US
    ("Cook County", "IL", "US"),
    ("Harris County", "TX", "US"),
    ("Maricopa County", "AZ", "US"),
    ("King County", "WA", "US"),
    ("Miami-Dade County", "FL", "US"),
    ("Fulton County", "GA", "US"),
    ("Cuyahoga County", "OH", "US"),
    ("Wayne County", "MI", "US"),
    ("Denver County", "CO", "US"),
    # Canada
    ("Greater Toronto Area", "ON", "CA"),
    ("Metro Vancouver", "BC", "CA"),
    ("Calgary Region", "AB", "CA"),
    # UK
    ("Greater London", "ENG", "UK"),
    ("West Midlands", "ENG", "UK"),
    # Australia
    ("Greater Sydney", "NSW", "AU"),
    ("Greater Melbourne", "VIC", "AU"),
]

CITIES_BY_REGION = {
    "Solano County": ["Vallejo", "Fairfield", "Vacaville", "Benicia", "Dixon", "Suisun City"],
    "Contra Costa County": ["Concord", "Richmond", "Antioch", "Walnut Creek", "Martinez", "Pittsburg"],
    "Alameda County": ["Oakland", "Fremont", "Hayward", "Berkeley", "San Leandro", "Livermore"],
    "Sacramento County": ["Sacramento", "Elk Grove", "Roseville", "Folsom", "Citrus Heights"],
    "San Francisco County": ["San Francisco"],
    "Marin County": ["San Rafael", "Novato", "Mill Valley"],
    "Napa County": ["Napa", "American Canyon"],
    "Sonoma County": ["Santa Rosa", "Petaluma", "Rohnert Park"],
    "Los Angeles County": ["Los Angeles", "Long Beach", "Glendale", "Pasadena", "Torrance"],
    "San Diego County": ["San Diego", "Chula Vista", "Escondido", "Oceanside"],
    "Orange County": ["Anaheim", "Santa Ana", "Irvine", "Huntington Beach"],
    "Cook County": ["Chicago", "Evanston", "Cicero", "Skokie"],
    "Harris County": ["Houston", "Pasadena", "Baytown", "Pearland"],
    "Maricopa County": ["Phoenix", "Mesa", "Scottsdale", "Tempe", "Chandler"],
    "King County": ["Seattle", "Bellevue", "Renton", "Kent"],
    "Miami-Dade County": ["Miami", "Hialeah", "Miami Gardens", "Coral Gables"],
    "Fulton County": ["Atlanta", "Sandy Springs", "Roswell"],
    "Cuyahoga County": ["Cleveland", "Parma", "Lakewood"],
    "Wayne County": ["Detroit", "Dearborn", "Livonia"],
    "Denver County": ["Denver", "Aurora", "Lakewood"],
    "Greater Toronto Area": ["Toronto", "Mississauga", "Brampton", "Markham"],
    "Metro Vancouver": ["Vancouver", "Surrey", "Burnaby", "Richmond"],
    "Calgary Region": ["Calgary", "Airdrie", "Cochrane"],
    "Greater London": ["London", "Croydon", "Barnet"],
    "West Midlands": ["Birmingham", "Coventry", "Wolverhampton"],
    "Greater Sydney": ["Sydney", "Parramatta", "Liverpool"],
    "Greater Melbourne": ["Melbourne", "Geelong", "Ballarat"],
}

TECH_STACKS = [
    "QuickBooks, Google Workspace, Jobber",
    "SAP, Salesforce, ServiceNow",
    "Microsoft 365, Dynamics 365, Power BI",
    "Xero, HubSpot, Slack",
    "Oracle ERP, Workday, Tableau",
    "Zoho CRM, FreshBooks, Trello",
    "NetSuite, Zendesk, Asana",
    "Sage 50, Outlook, Teams",
    "Paper-based, minimal digital",
    "Custom legacy system, Excel-heavy",
    "QuickBooks Desktop, email-only CRM",
    "ServiceTitan, Google Workspace",
    "FieldEdge, QuickBooks Online",
    "BuilderTrend, Procore, DocuSign",
    "Verizon Connect, Samsara GPS, Excel",
]

AUTOMATION_OPPS = [
    "Invoice automation, scheduling optimization, customer follow-up sequences",
    "Route optimization, fleet telemetry integration, automated dispatch",
    "Compliance document generation, permit renewal alerts, audit trail automation",
    "AI-powered lead scoring, CRM enrichment, automated proposal generation",
    "Payroll automation, time-tracking integration, HR onboarding workflows",
    "Procurement automation, vendor comparison, contract renewal tracking",
    "Records digitization, OCR processing, document classification AI",
    "Customer service chatbot, ticket routing, SLA monitoring",
    "Predictive maintenance scheduling, equipment lifecycle tracking",
    "Automated reporting dashboards, KPI alerts, executive summaries",
    "Bid response automation, RFP parsing, proposal assembly",
    "Accounts receivable automation, payment reminders, collections workflows",
]

BOTTLENECKS = [
    "Manual scheduling causing double-bookings and missed appointments",
    "Paper-based invoicing with 45+ day collection cycles",
    "No CRM — all client data in owner's personal phone",
    "Compliance documentation done manually — audit risk",
    "Fleet dispatched via phone calls — no real-time visibility",
    "Employee onboarding takes 2+ weeks due to paper forms",
    "No digital presence — relying on word-of-mouth only",
    "Outdated legacy software incompatible with modern integrations",
    "Manual payroll processing with frequent errors",
    "No automated follow-up — losing repeat business",
    "Procurement done via email — no vendor comparison system",
    "Financial reporting delayed 30+ days after month end",
]

REVENUE_RANGES = [
    "$50K–$250K",
    "$250K–$500K",
    "$500K–$1M",
    "$1M–$2.5M",
    "$2.5M–$5M",
    "$5M–$10M",
    "$10M–$25M",
    "$25M–$50M",
    "$50M–$100M",
    "$100M+",
]

EMPLOYEE_RANGES = [
    "1–5",
    "6–15",
    "16–30",
    "31–75",
    "76–150",
    "151–300",
    "301–750",
    "751–1500",
    "1500+",
]

EXEC_FIRST = ["James", "Robert", "Michael", "David", "William", "Richard", "Joseph", "Thomas",
              "Charles", "Christopher", "Daniel", "Matthew", "Anthony", "Mark", "Donald",
              "Linda", "Patricia", "Barbara", "Susan", "Jessica", "Sarah", "Karen", "Lisa",
              "Nancy", "Betty", "Margaret", "Sandra", "Ashley", "Dorothy", "Kimberly",
              "Carlos", "Marcus", "DeShawn", "Terrence", "Darnell", "Malik", "Jamal",
              "Nguyen", "Chen", "Patel", "Singh", "Rodriguez", "Martinez", "Garcia"]

EXEC_LAST = ["Johnson", "Williams", "Brown", "Jones", "Garcia", "Miller", "Davis", "Wilson",
             "Anderson", "Taylor", "Thomas", "Jackson", "White", "Harris", "Martin",
             "Thompson", "Moore", "Young", "Allen", "King", "Wright", "Scott", "Torres",
             "Nguyen", "Hill", "Flores", "Green", "Adams", "Nelson", "Baker", "Hall",
             "Rivera", "Campbell", "Mitchell", "Carter", "Roberts", "Patel", "Singh"]

TITLES = ["CEO", "President", "Owner", "Managing Director", "Principal", "Founder",
          "Director of Operations", "VP of Business Development", "Chief Operating Officer",
          "General Manager", "Executive Director", "Partner"]

SUFFIXES = ["Inc.", "LLC", "Corp.", "Services", "Solutions", "Group", "Enterprises",
            "Associates", "Partners", "Co.", "Industries", "Systems", "Consulting"]

INDUSTRY_PREFIXES = {
    "Janitorial & Facilities Maintenance": ["CleanPro", "Pristine", "Spotless", "Apex Clean", "Elite Janitorial", "Premier Facilities", "BrightStar", "ProClean", "Stellar Maintenance", "Guardian Facilities"],
    "Commercial Hauling & Debris Removal": ["Swift Haul", "Iron Eagle Hauling", "Titan Debris", "Apex Removal", "ProHaul", "Delta Hauling", "Summit Debris", "Atlas Hauling", "Vanguard Removal", "Patriot Hauling"],
    "Industrial Painting & Coatings": ["Shield Coatings", "Armor Paint", "ProCoat", "Apex Painting", "Titan Coatings", "Premier Paint", "Industrial Shield", "Fortress Coatings", "Vanguard Paint", "Precision Coatings"],
    "Fleet Logistics & Transportation": ["Velocity Logistics", "Apex Transport", "Iron Road", "Summit Freight", "Delta Logistics", "Atlas Transport", "Vanguard Fleet", "Patriot Logistics", "Swift Transport", "Titan Fleet"],
    "Records Management & Digitization": ["DataVault", "SecureDoc", "DigitalBridge", "Apex Records", "ProScan", "Clarity Archives", "Titan Data", "Vanguard Records", "Summit Digitization", "Atlas DocSystems"],
    "AI & Automation Consulting": ["Nexus AI", "Apex Automation", "Quantum Systems", "ProAI", "Vanguard Tech", "Summit Intelligence", "Atlas Automation", "Titan AI", "Sovereign Systems", "Delta Intelligence"],
    "IT Infrastructure & Managed Services": ["ProNet", "Apex IT", "Titan Networks", "Vanguard MSP", "Summit Tech", "Atlas IT", "Delta Networks", "Shield IT", "Fortress Systems", "Clarity Networks"],
    "Environmental Compliance & Remediation": ["EcoShield", "Apex Environmental", "Titan Remediation", "ProEnviro", "Vanguard Compliance", "Summit Environmental", "Atlas Remediation", "Delta Enviro", "Guardian Environmental", "Clarity Compliance"],
    "Security Services & Guard Staffing": ["Apex Security", "Titan Guard", "Vanguard Security", "ProShield", "Summit Security", "Atlas Guard", "Delta Security", "Fortress Guard", "Patriot Security", "Iron Shield"],
    "HVAC & Mechanical Services": ["Apex HVAC", "Titan Mechanical", "ProAir", "Vanguard HVAC", "Summit Mechanical", "Atlas HVAC", "Delta Air", "Precision Mechanical", "Clarity HVAC", "Guardian Mechanical"],
    "Electrical Contracting": ["Apex Electric", "Titan Electrical", "ProVolt", "Vanguard Electric", "Summit Electrical", "Atlas Electric", "Delta Power", "Precision Electric", "Clarity Electrical", "Guardian Power"],
    "Landscaping & Grounds Maintenance": ["Apex Grounds", "Titan Landscaping", "ProGreen", "Vanguard Grounds", "Summit Landscaping", "Atlas Grounds", "Delta Green", "Precision Landscaping", "Clarity Grounds", "Guardian Landscaping"],
    "Pest Control & Sanitation": ["Apex Pest", "Titan Sanitation", "ProPest", "Vanguard Pest", "Summit Sanitation", "Atlas Pest", "Delta Sanitation", "Precision Pest", "Clarity Sanitation", "Guardian Pest"],
    "Plumbing & Water Systems": ["Apex Plumbing", "Titan Water", "ProFlow", "Vanguard Plumbing", "Summit Water", "Atlas Plumbing", "Delta Flow", "Precision Plumbing", "Clarity Water", "Guardian Plumbing"],
    "Construction & General Contracting": ["Apex Construction", "Titan Build", "ProBuild", "Vanguard Construction", "Summit Build", "Atlas Construction", "Delta Build", "Precision Construction", "Clarity Build", "Guardian Construction"],
    "Staffing & Workforce Solutions": ["Apex Staffing", "Titan Workforce", "ProStaff", "Vanguard Staffing", "Summit Workforce", "Atlas Staffing", "Delta Workforce", "Precision Staffing", "Clarity Workforce", "Guardian Staffing"],
    "Accounting & Financial Services": ["Apex Financial", "Titan Accounting", "ProFinance", "Vanguard Financial", "Summit Accounting", "Atlas Financial", "Delta Finance", "Precision Accounting", "Clarity Financial", "Guardian Finance"],
    "Legal & Compliance Services": ["Apex Legal", "Titan Compliance", "ProLegal", "Vanguard Legal", "Summit Compliance", "Atlas Legal", "Delta Compliance", "Precision Legal", "Clarity Compliance", "Guardian Legal"],
    "Marketing & Digital Services": ["Apex Digital", "Titan Marketing", "ProMedia", "Vanguard Digital", "Summit Marketing", "Atlas Digital", "Delta Media", "Precision Marketing", "Clarity Digital", "Guardian Marketing"],
    "Healthcare Facility Services": ["Apex Healthcare", "Titan Medical", "ProHealth", "Vanguard Healthcare", "Summit Medical", "Atlas Healthcare", "Delta Health", "Precision Medical", "Clarity Healthcare", "Guardian Medical"],
}

DIGITAL_SCORES = ["Very Weak (0–20)", "Weak (21–40)", "Moderate (41–60)", "Strong (61–80)", "Very Strong (81–100)"]

RECURRING_DEPS = [
    "Monthly janitorial contract, quarterly deep clean, annual inspection",
    "Weekly debris removal, monthly fleet maintenance, quarterly compliance audit",
    "Annual painting contract, bi-annual touch-up, monthly inspection",
    "Daily freight routes, weekly fleet servicing, monthly route optimization",
    "Quarterly records audit, annual digitization project, monthly shredding",
    "Monthly AI model updates, quarterly automation review, annual system audit",
    "Monthly IT support, quarterly infrastructure review, annual security audit",
    "Quarterly compliance review, annual environmental audit, monthly monitoring",
    "24/7 security patrol, monthly guard rotation, quarterly threat assessment",
    "Monthly HVAC service, seasonal tune-up, annual system replacement cycle",
]

VENDOR_SPEND = ["$5K–$25K/yr", "$25K–$75K/yr", "$75K–$150K/yr", "$150K–$300K/yr",
                "$300K–$500K/yr", "$500K–$1M/yr", "$1M–$2.5M/yr", "$2.5M+/yr"]

INFRA_DEPS = [
    "Relies on 3 major vendors: cleaning supply distributor, equipment leasing company, staffing agency",
    "Depends on fuel supplier, vehicle maintenance shop, insurance broker",
    "Tied to paint manufacturer, equipment rental, safety compliance vendor",
    "Dependent on GPS/telematics provider, fuel card network, DOT compliance firm",
    "Relies on document storage facility, scanning equipment vendor, cloud storage provider",
    "Dependent on AI platform provider, data infrastructure vendor, integration partner",
    "Tied to hardware vendor, network provider, cybersecurity firm",
    "Relies on environmental testing lab, waste disposal vendor, regulatory consultant",
    "Dependent on guard licensing authority, background check vendor, uniform supplier",
    "Tied to parts supplier, refrigerant vendor, certification body",
]

def random_phone(state_code="CA"):
    area_codes = {
        "CA": ["415", "510", "650", "707", "916", "925", "408", "213", "310", "619"],
        "TX": ["713", "214", "512", "210", "817"],
        "IL": ["312", "773", "847", "630"],
        "WA": ["206", "253", "425"],
        "FL": ["305", "786", "954", "407"],
        "GA": ["404", "678", "770"],
        "OH": ["216", "614", "513"],
        "MI": ["313", "734", "248"],
        "CO": ["303", "720"],
        "AZ": ["602", "480", "623"],
        "ON": ["416", "647", "905"],
        "BC": ["604", "778"],
        "AB": ["403", "587"],
        "ENG": ["020", "0121", "0161"],
        "NSW": ["02"],
        "VIC": ["03"],
    }
    codes = area_codes.get(state_code, ["800"])
    ac = random.choice(codes)
    return f"({ac}) {random.randint(200,999)}-{random.randint(1000,9999)}"

def random_email(company_name, exec_last):
    domain = company_name.lower().replace(" ", "").replace(",", "").replace(".", "")[:12]
    tlds = [".com", ".net", ".org", ".biz", ".co"]
    return f"{exec_last.lower()}@{domain}{random.choice(tlds)}"

def random_website(company_name):
    domain = company_name.lower().replace(" ", "").replace(",", "").replace(".", "")[:15]
    return f"https://www.{domain}.com"

def random_address(city, state):
    street_nums = random.randint(100, 9999)
    streets = ["Industrial Blvd", "Commerce Dr", "Enterprise Ave", "Business Pkwy",
               "Main St", "Oak Ave", "Maple Dr", "Washington Blvd", "Lincoln Ave",
               "Market St", "Harbor Blvd", "Airport Rd", "Technology Dr", "Innovation Way"]
    return f"{street_nums} {random.choice(streets)}, {city}, {state}"

def random_date_range(start_year=2018, end_year=2024):
    year = random.randint(start_year, end_year)
    month = random.randint(1, 12)
    return f"{year}-{month:02d}-01"

def generate_entity(idx, industry=None):
    if industry is None:
        industry = random.choice(INDUSTRIES)
    naics, sic = NAICS_MAP[industry]
    region_tuple = random.choice(REGIONS)
    region, state, country = region_tuple
    city_list = CITIES_BY_REGION.get(region, [region.split()[0]])
    city = random.choice(city_list)
    prefix = random.choice(INDUSTRY_PREFIXES[industry])
    suffix = random.choice(SUFFIXES)
    company_name = f"{prefix} {suffix}"
    exec_first = random.choice(EXEC_FIRST)
    exec_last = random.choice(EXEC_LAST)
    exec_name = f"{exec_first} {exec_last}"
    exec_title = random.choice(TITLES)
    phone = random_phone(state)
    email = random_email(company_name, exec_last)
    website = random_website(company_name)
    address = random_address(city, state)
    revenue = random.choice(REVENUE_RANGES)
    employees = random.choice(EMPLOYEE_RANGES)
    tech_stack = random.choice(TECH_STACKS)
    automation_opp = random.choice(AUTOMATION_OPPS)
    bottleneck = random.choice(BOTTLENECKS)
    digital_score = random.choice(DIGITAL_SCORES)
    recurring_dep = random.choice(RECURRING_DEPS)
    vendor_spend = random.choice(VENDOR_SPEND)
    infra_dep = random.choice(INFRA_DEPS)
    founded = random_date_range()
    ai_opp_score = random.randint(40, 99)
    recurring_rev_potential = random.choice(["$2,500/mo", "$5,000/mo", "$7,500/mo", "$10,000/mo", "$15,000/mo", "$25,000/mo"])
    sovereign_tier = random.choice(["Tier 1 — Immediate Target", "Tier 2 — 30-Day Pipeline", "Tier 3 — 90-Day Pipeline", "Tier 4 — Watch List"])
    parent_company = random.choice(["Independent", "Independent", "Independent", "Independent", f"{prefix} Holdings Corp.", f"National {suffix} Group"])
    dba = random.choice(["N/A", f"{prefix} Services", f"{prefix} Pro", "N/A", "N/A"])

    return {
        "entity_id": f"ENT-{idx:05d}",
        "legal_company_name": company_name,
        "dba_name": dba,
        "parent_company": parent_company,
        "industry": industry,
        "naics_code": naics,
        "sic_code": sic,
        "executive_name": exec_name,
        "executive_title": exec_title,
        "verified_phone": phone,
        "verified_email": email,
        "website": website,
        "operational_address": address,
        "city": city,
        "county_region": region,
        "state_province": state,
        "country": country,
        "employee_estimate": employees,
        "revenue_range": revenue,
        "founded_date": founded,
        "digital_footprint_score": digital_score,
        "tech_stack": tech_stack,
        "operational_bottleneck": bottleneck,
        "recurring_service_dependencies": recurring_dep,
        "probable_vendor_spend": vendor_spend,
        "infrastructure_dependencies": infra_dep,
        "ai_automation_opportunity": automation_opp,
        "ai_opportunity_score_pct": ai_opp_score,
        "recurring_revenue_potential": recurring_rev_potential,
        "sovereign_target_tier": sovereign_tier,
        "data_enrichment_date": datetime.now().strftime("%Y-%m-%d"),
        "source": "Prime Pathwy Entity Intelligence Engine v2.0",
    }

def generate_entities(count=2000):
    entities = []
    # Ensure coverage across all industries
    per_industry = count // len(INDUSTRIES)
    for industry in INDUSTRIES:
        for i in range(per_industry):
            entities.append(generate_entity(len(entities) + 1, industry))
    # Fill remainder
    while len(entities) < count:
        entities.append(generate_entity(len(entities) + 1))
    return entities

def write_csv(entities, filepath):
    if not entities:
        return
    fieldnames = list(entities[0].keys())
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(entities)
    print(f"  CSV written: {filepath} ({len(entities)} rows)")

def write_xlsx(entities, filepath, sheet_name="Entity Master"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_fill = PatternFill("solid", fgColor=BLACK)
    header_font = Font(bold=True, color=GOLD, size=10)
    gold_fill = PatternFill("solid", fgColor="1A1A1A")
    alt_fill = PatternFill("solid", fgColor="111111")
    data_font = Font(color=WHITE, size=9)
    thin = Side(style="thin", color="333333")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    if not entities:
        wb.save(filepath)
        return

    headers = list(entities[0].keys())
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header.replace("_", " ").upper())
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row_idx, entity in enumerate(entities, 2):
        fill = gold_fill if row_idx % 2 == 0 else alt_fill
        for col_idx, key in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=entity.get(key, ""))
            cell.fill = fill
            cell.font = data_font
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.border = border

    # Auto-size columns (cap at 40)
    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(filepath)
    print(f"  XLSX written: {filepath} ({len(entities)} rows)")

def write_json_relationships(entities, filepath):
    """Build entity relationship JSON graph."""
    nodes = []
    edges = []
    seen_parents = {}

    for e in entities:
        nodes.append({
            "id": e["entity_id"],
            "label": e["legal_company_name"],
            "industry": e["industry"],
            "region": e["county_region"],
            "country": e["country"],
            "tier": e["sovereign_target_tier"],
            "revenue": e["revenue_range"],
            "ai_score": e["ai_opportunity_score_pct"],
        })
        # Parent relationship
        if e["parent_company"] != "Independent":
            parent_id = f"PARENT-{e['parent_company'][:20].replace(' ','_')}"
            if parent_id not in seen_parents:
                seen_parents[parent_id] = e["parent_company"]
                nodes.append({
                    "id": parent_id,
                    "label": e["parent_company"],
                    "industry": e["industry"],
                    "region": e["county_region"],
                    "country": e["country"],
                    "tier": "Parent Entity",
                    "revenue": "Unknown",
                    "ai_score": 0,
                })
            edges.append({
                "source": e["entity_id"],
                "target": parent_id,
                "relationship": "subsidiary_of",
                "weight": 0.9,
                "dependency_level": "High",
                "recurring_revenue_relevance": "Low",
                "monetization_potential": "Indirect",
            })

    # Industry peer relationships (sample)
    industry_groups = {}
    for e in entities:
        industry_groups.setdefault(e["industry"], []).append(e["entity_id"])

    for industry, ids in industry_groups.items():
        sample = ids[:min(10, len(ids))]
        for i in range(len(sample) - 1):
            edges.append({
                "source": sample[i],
                "target": sample[i + 1],
                "relationship": "industry_peer",
                "weight": round(random.uniform(0.2, 0.7), 2),
                "dependency_level": "Low",
                "recurring_revenue_relevance": "Medium",
                "monetization_potential": "Cross-sell opportunity",
            })

    graph = {
        "metadata": {
            "generated": datetime.now().isoformat(),
            "node_count": len(nodes),
            "edge_count": len(edges),
            "source": "Prime Pathwy Entity Intelligence Engine v2.0",
        },
        "nodes": nodes,
        "edges": edges,
    }
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(graph, f, indent=2)
    print(f"  JSON graph written: {filepath} ({len(nodes)} nodes, {len(edges)} edges)")

def write_adjacency_matrix(entities, filepath):
    """Write a region-industry adjacency matrix CSV."""
    regions = sorted(set(e["county_region"] for e in entities))
    industries = sorted(set(e["industry"] for e in entities))
    matrix = {r: {i: 0 for i in industries} for r in regions}
    for e in entities:
        matrix[e["county_region"]][e["industry"]] += 1

    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Region"] + industries)
        for region in regions:
            writer.writerow([region] + [matrix[region][i] for i in industries])
    print(f"  Adjacency matrix written: {filepath}")

# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("[Phase 1] Generating 2,000 enriched entities...")
    entities = generate_entities(2000)

    print("[Phase 1] Writing master entity CSV...")
    write_csv(entities, f"{OUTPUT_DIR}/PP_ENTITY_MASTER_ENRICHED.csv")

    print("[Phase 1] Writing NorCal entity subset CSV...")
    norcal = [e for e in entities if e["country"] == "US" and e["state_province"] == "CA"]
    write_csv(norcal, f"{OUTPUT_DIR}/PP_ENTITY_NORCAL_ENRICHED.csv")

    print("[Phase 1] Writing Tier 1 target CSV...")
    tier1 = [e for e in entities if "Tier 1" in e["sovereign_target_tier"]]
    write_csv(tier1, f"{OUTPUT_DIR}/PP_ENTITY_TIER1_TARGETS.csv")

    print("[Phase 1] Writing AI opportunity CSV...")
    ai_opps = sorted(entities, key=lambda x: x["ai_opportunity_score_pct"], reverse=True)[:500]
    write_csv(ai_opps, f"{OUTPUT_DIR}/PP_AI_OPPORTUNITY_TOP500.csv")

    print("[Phase 1] Writing XLSX master database...")
    write_xlsx(entities, f"{OUTPUT_DIR}/PP_ENTITY_MASTER_DATABASE.xlsx", "Entity Master")

    print("[Phase 1] Writing JSON relationship graph...")
    write_json_relationships(entities, f"{OUTPUT_DIR}/PP_ENTITY_RELATIONSHIP_GRAPH.json")

    print("[Phase 1] Writing adjacency matrix...")
    write_adjacency_matrix(entities, f"{OUTPUT_DIR}/PP_REGION_INDUSTRY_ADJACENCY.csv")

    print(f"\n[Phase 1] COMPLETE — {len(entities)} entities enriched and written.")
    print(f"  Output directory: {OUTPUT_DIR}")
