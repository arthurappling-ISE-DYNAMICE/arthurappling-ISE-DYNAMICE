#!/usr/bin/env python3
"""
TOOL_GENERATE_XLSX_V1.py
Prime Pathwy Sovereign Intelligence Engine
Generates formatted XLSX databases from all CSV datasets.
Author: Arthur F. Appling Sr.
Version: 1.1.0
"""

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Color constants: Matte Black and Gold
MATTE_BLACK = "0B0B0B"
GOLD = "C9A646"
DARK_GRAY = "1A1A1A"
LIGHT_GRAY = "2C2C2C"
WHITE = "FFFFFF"

CSV_SOURCES = [
    {
        "csv_path": "/home/ubuntu/prime-pathwy-sovereign-vault/vault/phase8_datasets/csv/procurement_intelligence.csv",
        "xlsx_path": "/home/ubuntu/prime-pathwy-sovereign-vault/vault/phase8_datasets/xlsx/procurement_intelligence.xlsx",
        "sheet_name": "Procurement Intelligence"
    },
    {
        "csv_path": "/home/ubuntu/prime-pathwy-sovereign-vault/vault/phase8_datasets/csv/sovereign_system_pricing_tiers.csv",
        "xlsx_path": "/home/ubuntu/prime-pathwy-sovereign-vault/vault/phase8_datasets/xlsx/sovereign_system_pricing_tiers.xlsx",
        "sheet_name": "Pricing Tiers"
    },
    {
        "csv_path": "/home/ubuntu/prime-pathwy-sovereign-vault/vault/phase8_datasets/csv/kpi_sla_performance_template.csv",
        "xlsx_path": "/home/ubuntu/prime-pathwy-sovereign-vault/vault/phase8_datasets/xlsx/kpi_sla_performance_template.xlsx",
        "sheet_name": "KPI & SLA Targets"
    },
    {
        "csv_path": "/home/ubuntu/prime-pathwy-sovereign-vault/vault/phase7_knowledge_graph/csv/entity_adjacency_matrix.csv",
        "xlsx_path": "/home/ubuntu/prime-pathwy-sovereign-vault/vault/phase8_datasets/xlsx/entity_adjacency_matrix.xlsx",
        "sheet_name": "Entity Adjacency Matrix"
    },
    {
        "csv_path": "/home/ubuntu/prime-pathwy-sovereign-vault/vault/phase7_knowledge_graph/csv/monetization_opportunity_map.csv",
        "xlsx_path": "/home/ubuntu/prime-pathwy-sovereign-vault/vault/phase8_datasets/xlsx/monetization_opportunity_map.xlsx",
        "sheet_name": "Monetization Opportunities"
    }
]

def apply_header_style(ws, row_num, col_count):
    """Apply Matte Black + Gold header styling to a worksheet row."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = PatternFill(start_color=MATTE_BLACK, end_color=MATTE_BLACK, fill_type="solid")
        cell.font = Font(color=GOLD, bold=True, name="Calibri", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            bottom=Side(style="medium", color=GOLD),
            right=Side(style="thin", color=DARK_GRAY)
        )

def apply_data_row_style(ws, row_num, col_count, is_even):
    """Apply alternating row styling to data rows."""
    bg_color = DARK_GRAY if is_even else LIGHT_GRAY
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        cell.font = Font(color=WHITE, name="Calibri", size=10)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

def csv_to_xlsx(csv_path, xlsx_path, sheet_name):
    """Convert a CSV file to a formatted XLSX file."""
    df = pd.read_csv(csv_path)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Set tab color to gold
    ws.sheet_properties.tabColor = GOLD

    # Write header row
    headers = list(df.columns)
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header.replace("_", " ").upper())

    apply_header_style(ws, 1, len(headers))

    # Write data rows
    for row_num, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
        apply_data_row_style(ws, row_num, len(headers), row_num % 2 == 0)

    # Auto-size columns
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 4, 50)

    # Freeze the header row
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

    wb.save(xlsx_path)
    print(f"[SUCCESS] Generated: {xlsx_path}")

def main():
    os.makedirs("/home/ubuntu/prime-pathwy-sovereign-vault/vault/phase8_datasets/xlsx", exist_ok=True)

    for source in CSV_SOURCES:
        if os.path.exists(source["csv_path"]):
            csv_to_xlsx(source["csv_path"], source["xlsx_path"], source["sheet_name"])
        else:
            print(f"[WARNING] CSV not found: {source['csv_path']}")

    print("\n[COMPLETE] All XLSX databases generated successfully.")

if __name__ == "__main__":
    main()
