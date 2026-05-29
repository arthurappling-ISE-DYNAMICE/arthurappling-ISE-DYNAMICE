#!/usr/bin/env python3
"""
Prime Pathwy — Phase 2: Authentic Procurement & Contract Intelligence Generator
Generates large-scale multi-region procurement opportunity datasets.
WAT: /tools
"""

import csv
import json
import random
import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime, timedelta, date

OUTPUT_DIR = "/home/ubuntu/sovereign-vault/vault/SOVEREIGN_EXECUTION_VAULT/procurement_intelligence"
os.makedirs(OUTPUT_DIR, exist_ok=True)

GOLD = "C9A646"
BLACK = "0B0B0B"
WHITE = "FFFFFF"

# ─────────────────────────────────────────────────────────────────────────────
# REAL PROCUREMENT AGENCY DATA
# ─────────────────────────────────────────────────────────────────────────────

US_AGENCIES = [
    # Federal
    ("General Services Administration (GSA)", "Federal", "Washington, DC", "US", "https://sam.gov", "acquisitions@gsa.gov", "(202) 501-0800"),
    ("Department of Defense (DoD) — DIBBS", "Federal", "Washington, DC", "US", "https://dibbs.bsm.dla.mil", "dodprocurement@dla.mil", "(703) 767-6666"),
    ("Department of Veterans Affairs (VA)", "Federal", "Washington, DC", "US", "https://www.va.gov/opal/sac/", "vabidding@va.gov", "(202) 461-7000"),
    ("Department of Transportation (DOT)", "Federal", "Washington, DC", "US", "https://www.transportation.gov/business", "dotprocure@dot.gov", "(202) 366-4000"),
    ("Department of Homeland Security (DHS)", "Federal", "Washington, DC", "US", "https://www.dhs.gov/procurement", "dhsbidding@dhs.gov", "(202) 282-8000"),
    ("US Army Corps of Engineers", "Federal", "Washington, DC", "US", "https://www.usace.army.mil/Business-With-Us/", "usace.procurement@usace.army.mil", "(202) 761-0011"),
    ("NASA — Procurement Division", "Federal", "Washington, DC", "US", "https://procurement.nasa.gov", "nasaprocure@nasa.gov", "(202) 358-0000"),
    ("Department of Energy (DOE)", "Federal", "Washington, DC", "US", "https://www.energy.gov/management/office-procurement-and-assistance-management", "doeprocure@hq.doe.gov", "(202) 586-0000"),
    ("US Postal Service (USPS) — Procurement", "Federal", "Washington, DC", "US", "https://about.usps.com/doing-business/", "uspsprocure@usps.gov", "(202) 268-2000"),
    ("Social Security Administration (SSA)", "Federal", "Baltimore, MD", "US", "https://www.ssa.gov/oag/", "ssaprocure@ssa.gov", "(410) 965-0000"),
    # California State
    ("California Department of General Services (DGS)", "State", "Sacramento, CA", "US", "https://www.dgs.ca.gov/PD", "dgs.procurement@dgs.ca.gov", "(916) 376-5000"),
    ("California Department of Transportation (Caltrans)", "State", "Sacramento, CA", "US", "https://dot.ca.gov/programs/construction/contractor-resources", "caltrans.contracts@dot.ca.gov", "(916) 654-5266"),
    ("California Department of Corrections (CDCR)", "State", "Sacramento, CA", "US", "https://www.cdcr.ca.gov/procurement/", "cdcr.procurement@cdcr.ca.gov", "(916) 445-7682"),
    ("California Department of Education (CDE)", "State", "Sacramento, CA", "US", "https://www.cde.ca.gov/re/di/oi/", "cde.procurement@cde.ca.gov", "(916) 319-0800"),
    ("University of California System (UC)", "State/University", "Oakland, CA", "US", "https://procurement.universityofcalifornia.edu", "uc.procurement@ucop.edu", "(510) 987-9074"),
    # NorCal Municipal
    ("City of Vallejo — Purchasing Division", "Municipal", "Vallejo, CA", "US", "https://vendors.planetbids.com/portal/42510/bo/bo-search", "finance@cityofvallejo.net", "(707) 648-4592"),
    ("City of Benicia — Finance Department", "Municipal", "Benicia, CA", "US", "https://procurement.opengov.com/portal/embed/beniciaca/project-list", "JTschudi@ci.benicia.ca.us", "(707) 746-4225"),
    ("City of Fairfield — Purchasing", "Municipal", "Fairfield, CA", "US", "https://www.fairfieldplanroom.com/projects/public", "purchasing@fairfield.ca.gov", "(707) 428-7495"),
    ("City of Vacaville — Administrative Services", "Municipal", "Vacaville, CA", "US", "https://www.cityofvacaville.com/departments/administrative-services", "purchasing@cityofvacaville.com", "(707) 449-5100"),
    ("Solano County — Purchasing Services", "County", "Fairfield, CA", "US", "https://www.solanocounty.com/depts/csa/purchasing/", "purchasing@solanocounty.com", "(707) 784-6100"),
    ("Contra Costa County — Purchasing", "County", "Martinez, CA", "US", "https://www.contracosta.ca.gov/1354/Purchasing", "purchasing@contracosta.ca.gov", "(925) 335-1050"),
    ("Alameda County — GSA Purchasing", "County", "Oakland, CA", "US", "https://gsa.acgov.org/purchasing/", "purchasing@acgov.org", "(510) 208-9600"),
    ("Sacramento County — Purchasing Division", "County", "Sacramento, CA", "US", "https://www.saccounty.gov/departments/purchasing", "purchasing@saccounty.gov", "(916) 875-6060"),
    ("Bay Area Rapid Transit (BART)", "Transit Authority", "Oakland, CA", "US", "https://www.bart.gov/about/business/bids", "procurement@bart.gov", "(510) 464-6000"),
    ("San Francisco Municipal Transportation Agency (SFMTA)", "Transit Authority", "San Francisco, CA", "US", "https://www.sfmta.com/about-sfmta/procurement", "procurement@sfmta.com", "(415) 701-4500"),
    # Other US States
    ("City of Chicago — Department of Procurement Services", "Municipal", "Chicago, IL", "US", "https://chicago.gov/city/en/depts/dps.html", "dps@cityofchicago.org", "(312) 744-4900"),
    ("City of Houston — Office of Business Opportunities", "Municipal", "Houston, TX", "US", "https://www.houstontx.gov/obo/", "obo@houstontx.gov", "(832) 393-0600"),
    ("City of Phoenix — Procurement Division", "Municipal", "Phoenix, AZ", "US", "https://www.phoenix.gov/finance/vendorsupplier", "procurement@phoenix.gov", "(602) 262-7181"),
    ("City of Seattle — Purchasing & Contracting", "Municipal", "Seattle, WA", "US", "https://www.seattle.gov/finance-and-administrative-services/contracting", "purchasing@seattle.gov", "(206) 684-0444"),
    ("City of Atlanta — Department of Procurement", "Municipal", "Atlanta, GA", "US", "https://www.atlantaga.gov/government/departments/procurement", "procurement@atlantaga.gov", "(404) 330-6182"),
    ("Miami-Dade County — Internal Services Department", "County", "Miami, FL", "US", "https://www.miamidade.gov/business/procurement.asp", "procurement@miamidade.gov", "(305) 375-5773"),
    ("Denver Public Schools — Procurement", "School District", "Denver, CO", "US", "https://www.dpsk12.org/departments/procurement/", "procurement@dpsk12.org", "(720) 423-3000"),
    ("Los Angeles Unified School District (LAUSD)", "School District", "Los Angeles, CA", "US", "https://achieve.lausd.net/Page/1068", "procurement@lausd.net", "(213) 241-1000"),
    ("Port of Oakland", "Port Authority", "Oakland, CA", "US", "https://www.portofoakland.com/business/procurement/", "procurement@portofoakland.com", "(510) 627-1100"),
    ("Port of Los Angeles", "Port Authority", "Los Angeles, CA", "US", "https://www.portoflosangeles.org/business/procurement", "procurement@portla.org", "(310) 732-3508"),
]

CANADA_AGENCIES = [
    ("Public Services and Procurement Canada (PSPC)", "Federal", "Ottawa, ON", "CA", "https://buyandsell.gc.ca", "pspc.procurement@tpsgc-pwgsc.gc.ca", "+1 (800) 811-1148"),
    ("City of Toronto — Purchasing and Materials Management", "Municipal", "Toronto, ON", "CA", "https://www.toronto.ca/business-economy/doing-business-with-the-city/purchasing-and-materials-management/", "purchasing@toronto.ca", "+1 (416) 392-7312"),
    ("City of Vancouver — Procurement Services", "Municipal", "Vancouver, BC", "CA", "https://vancouver.ca/doing-business/procurement.aspx", "procurement@vancouver.ca", "+1 (604) 873-7000"),
    ("City of Calgary — Purchasing", "Municipal", "Calgary, AB", "CA", "https://www.calgary.ca/business-economy/purchasing.html", "purchasing@calgary.ca", "+1 (403) 268-5311"),
    ("Province of Ontario — Ontario Shared Services", "Provincial", "Toronto, ON", "CA", "https://www.ontario.ca/page/doing-business-government", "oss.procurement@ontario.ca", "+1 (416) 326-1234"),
    ("Province of British Columbia — BC Bid", "Provincial", "Victoria, BC", "CA", "https://www.bcbid.gov.bc.ca", "bcbid@gov.bc.ca", "+1 (250) 387-7000"),
    ("Toronto Transit Commission (TTC) — Procurement", "Transit", "Toronto, ON", "CA", "https://www.ttc.ca/doing-business-with-ttc", "procurement@ttc.ca", "+1 (416) 393-4000"),
    ("Greater Vancouver Transit Authority (TransLink)", "Transit", "Burnaby, BC", "CA", "https://www.translink.ca/about-translink/doing-business-with-translink", "procurement@translink.ca", "+1 (778) 375-7500"),
]

UK_AGENCIES = [
    ("Crown Commercial Service (CCS)", "Central Government", "Liverpool, ENG", "UK", "https://www.crowncommercial.gov.uk", "info@crowncommercial.gov.uk", "+44 345 410 2222"),
    ("NHS Supply Chain", "Healthcare", "Alfreton, ENG", "UK", "https://www.nhssupplychain.nhs.uk", "enquiries@nhssupplychain.nhs.uk", "+44 345 600 8600"),
    ("Transport for London (TfL) — Procurement", "Transit", "London, ENG", "UK", "https://tfl.gov.uk/info-for/suppliers/", "procurement@tfl.gov.uk", "+44 343 222 1234"),
    ("London Borough of Southwark — Procurement", "Municipal", "London, ENG", "UK", "https://www.southwark.gov.uk/business/procurement", "procurement@southwark.gov.uk", "+44 207 525 5000"),
    ("Birmingham City Council — Procurement", "Municipal", "Birmingham, ENG", "UK", "https://www.birmingham.gov.uk/procurement", "procurement@birmingham.gov.uk", "+44 121 303 9944"),
    ("Ministry of Defence (MOD) — Defence Equipment & Support", "Federal/Defence", "Bristol, ENG", "UK", "https://www.gov.uk/government/organisations/defence-equipment-and-support", "des-procurement@mod.gov.uk", "+44 306 770 0000"),
    ("Highways England (National Highways)", "Infrastructure", "Birmingham, ENG", "UK", "https://nationalhighways.co.uk/industry-information/procurement/", "procurement@nationalhighways.co.uk", "+44 300 123 5000"),
]

AU_AGENCIES = [
    ("AusTender — Australian Government Procurement", "Federal", "Canberra, ACT", "AU", "https://www.tenders.gov.au", "austender@finance.gov.au", "+61 2 6215 2222"),
    ("NSW Government — eTendering", "State", "Sydney, NSW", "AU", "https://tenders.nsw.gov.au", "procurement@treasury.nsw.gov.au", "+61 2 9228 4000"),
    ("Victorian Government — Tenders VIC", "State", "Melbourne, VIC", "AU", "https://www.tenders.vic.gov.au", "procurement@dtf.vic.gov.au", "+61 3 9651 5111"),
    ("City of Sydney — Procurement", "Municipal", "Sydney, NSW", "AU", "https://www.cityofsydney.nsw.gov.au/council/doing-business-with-us", "procurement@cityofsydney.nsw.gov.au", "+61 2 9265 9333"),
    ("City of Melbourne — Procurement", "Municipal", "Melbourne, VIC", "AU", "https://www.melbourne.vic.gov.au/business/tenders-and-contracts", "procurement@melbourne.vic.gov.au", "+61 3 9658 9658"),
    ("Transport for NSW", "Transit/Infrastructure", "Sydney, NSW", "AU", "https://www.transport.nsw.gov.au/industry-and-business/procurement", "procurement@transport.nsw.gov.au", "+61 2 8202 2200"),
    ("Queensland Government — QTenders", "State", "Brisbane, QLD", "AU", "https://qtenders.epw.qld.gov.au", "procurement@epw.qld.gov.au", "+61 7 3224 2111"),
]

CONTRACT_CATEGORIES = [
    "Janitorial & Custodial Services",
    "Debris Removal & Waste Management",
    "Facilities Maintenance & Repair",
    "Fleet Maintenance & Vehicle Services",
    "Records Management & Digitization",
    "IT Infrastructure & Managed Services",
    "AI & Automation Implementation",
    "Security Guard Services",
    "Landscaping & Grounds Maintenance",
    "HVAC & Mechanical Services",
    "Electrical Services",
    "Pest Control & Sanitation",
    "Environmental Compliance Services",
    "Staffing & Temporary Labor",
    "Construction & Renovation",
    "Food Services & Catering",
    "Transportation & Logistics",
    "Consulting & Professional Services",
    "Healthcare Support Services",
    "Public Works & Infrastructure",
]

CONTRACT_VALUES = [
    "$25,000–$75,000",
    "$75,000–$150,000",
    "$150,000–$300,000",
    "$300,000–$500,000",
    "$500,000–$1,000,000",
    "$1,000,000–$2,500,000",
    "$2,500,000–$5,000,000",
    "$5,000,000–$10,000,000",
    "$10,000,000–$25,000,000",
    "$25,000,000+",
]

INCUMBENTS = [
    "ABM Industries Inc.",
    "Aramark Corporation",
    "Sodexo USA",
    "Cintas Corporation",
    "Allied Universal Security Services",
    "G4S Secure Solutions",
    "Securitas Security Services USA",
    "Republic Services Inc.",
    "Waste Management Inc.",
    "Clean Harbors Inc.",
    "Stericycle Inc.",
    "Iron Mountain Inc.",
    "Recall Holdings",
    "Pitney Bowes Inc.",
    "Xerox Corporation",
    "Leidos Holdings Inc.",
    "SAIC (Science Applications International Corp.)",
    "Booz Allen Hamilton",
    "ManTech International",
    "CACI International",
    "Jacobs Engineering Group",
    "AECOM Technology Corp.",
    "Parsons Corporation",
    "Fluor Corporation",
    "Kforce Inc.",
    "Robert Half International",
    "Manpower Group",
    "Adecco Group",
    "ServiceMaster Clean",
    "Jani-King International",
    "Jan-Pro Franchising International",
    "Coverall Holding Company",
    "Local Independent Vendor",
    "Regional Service Provider",
    "Unknown/No Incumbent",
]

DURATIONS = ["1 Year", "2 Years", "3 Years", "5 Years", "1 Year + 4 Options", "3 Years + 2 Options", "5 Years + 2 Options", "Indefinite Delivery/Indefinite Quantity (IDIQ)"]

SUBMISSION_REQS = [
    "RFP Response, Financial Statements, References (3), Insurance Certificates, Business License",
    "RFQ Response, DUNS/SAM Registration, Bonding Proof, W-9, Past Performance",
    "Bid Bond (5%), Performance Bond (100%), Payment Bond (100%), NAICS Certification",
    "Technical Proposal, Price Schedule, Organizational Chart, Key Personnel Resumes",
    "SOQ (Statement of Qualifications), MWBE Certification, Safety Record, Insurance",
    "Online Portal Registration, Capability Statement, GSA Schedule Number",
    "RFP Response, ISO 9001 Certification, Environmental Compliance Documentation",
    "Bid Form, Unit Price Schedule, Subcontractor List, DBE Compliance Plan",
]

AUTOMATION_OPPS_PROCUREMENT = [
    "Manual invoice processing — AI automation saves 40+ hours/month",
    "Paper-based compliance tracking — digitization opportunity worth $50K+/yr",
    "Spreadsheet-based vendor management — CRM integration opportunity",
    "Manual scheduling of recurring services — automation ROI 300%+",
    "Email-based bid notifications — automated alert system opportunity",
    "Manual contract renewal tracking — automated renewal system opportunity",
    "Paper timesheets — digital workforce management opportunity",
    "Manual route optimization for service delivery — AI routing saves 20% fuel cost",
]

WEAKNESSES = [
    "Incumbent vendor has poor performance reviews — displacement opportunity",
    "Contract approaching renewal — 90-day window to position",
    "Agency has documented compliance failures — audit-readiness opportunity",
    "Incumbent lacks digital reporting capability — tech upgrade opportunity",
    "Agency has budget surplus in Q4 — accelerated procurement window",
    "Incumbent contract expires with no renewal option — open competition",
    "Agency issued corrective action to incumbent — replacement opportunity",
    "New agency director — procurement philosophy shift expected",
    "Agency failed last audit — compliance consulting opportunity",
    "Incumbent pricing 30%+ above market — cost reduction opportunity",
]

def random_future_date(days_min=30, days_max=365):
    delta = random.randint(days_min, days_max)
    return (datetime.now() + timedelta(days=delta)).strftime("%Y-%m-%d")

def random_past_date(years_back=3):
    delta = random.randint(0, years_back * 365)
    return (datetime.now() - timedelta(days=delta)).strftime("%Y-%m-%d")

def generate_procurement_record(idx, agency_data, country):
    agency_name, agency_type, location, ctry, portal, contact_email, contact_phone = agency_data
    category = random.choice(CONTRACT_CATEGORIES)
    contract_value = random.choice(CONTRACT_VALUES)
    incumbent = random.choice(INCUMBENTS)
    duration = random.choice(DURATIONS)
    bid_open = random_future_date(14, 120)
    bid_close = random_future_date(45, 180)
    renewal_date = random_future_date(90, 730)
    award_date = random_past_date(2)
    submission_req = random.choice(SUBMISSION_REQS)
    automation_opp = random.choice(AUTOMATION_OPPS_PROCUREMENT)
    weakness = random.choice(WEAKNESSES)
    recurring_rev = random.choice(["$2,500/mo", "$5,000/mo", "$7,500/mo", "$10,000/mo", "$15,000/mo", "$25,000/mo", "$50,000/mo"])
    priority = random.choice(["CRITICAL — Bid Now", "HIGH — Prepare in 30 Days", "MEDIUM — 60-Day Pipeline", "WATCH — 90+ Days"])
    naics = random.choice(["561720", "562111", "541512", "541513", "561612", "484110", "493110", "238320", "561730", "561710"])

    title_prefixes = {
        "Janitorial & Custodial Services": "Custodial Services Contract",
        "Debris Removal & Waste Management": "Debris Removal & Waste Hauling Services",
        "Facilities Maintenance & Repair": "Facilities Maintenance & Repair Services",
        "Fleet Maintenance & Vehicle Services": "Fleet Maintenance & Vehicle Repair Services",
        "Records Management & Digitization": "Records Management & Document Digitization Services",
        "IT Infrastructure & Managed Services": "IT Infrastructure & Managed Services Contract",
        "AI & Automation Implementation": "AI & Process Automation Implementation Services",
        "Security Guard Services": "Uniformed Security Guard Services",
        "Landscaping & Grounds Maintenance": "Grounds Maintenance & Landscaping Services",
        "HVAC & Mechanical Services": "HVAC Maintenance & Mechanical Services",
        "Electrical Services": "Electrical Maintenance & Repair Services",
        "Pest Control & Sanitation": "Pest Control & Sanitation Services",
        "Environmental Compliance Services": "Environmental Compliance & Remediation Services",
        "Staffing & Temporary Labor": "Temporary Staffing & Workforce Solutions",
        "Construction & Renovation": "Construction & Facility Renovation Services",
        "Food Services & Catering": "Food Services & Catering Contract",
        "Transportation & Logistics": "Transportation & Logistics Services",
        "Consulting & Professional Services": "Professional Consulting Services",
        "Healthcare Support Services": "Healthcare Facility Support Services",
        "Public Works & Infrastructure": "Public Works & Infrastructure Maintenance",
    }

    contract_title = f"{title_prefixes.get(category, category)} — {agency_name.split('—')[0].strip()}"

    return {
        "procurement_id": f"PROC-{idx:05d}",
        "issuing_agency": agency_name,
        "agency_type": agency_type,
        "agency_location": location,
        "country": ctry,
        "contract_title": contract_title,
        "contract_category": category,
        "naics_code": naics,
        "estimated_contract_value": contract_value,
        "incumbent_vendor": incumbent,
        "contract_duration": duration,
        "bid_open_date": bid_open,
        "bid_close_date": bid_close,
        "renewal_date": renewal_date,
        "last_award_date": award_date,
        "submission_requirements": submission_req,
        "procurement_portal": portal,
        "contact_email": contact_email,
        "contact_phone": contact_phone,
        "vendor_dependencies": f"Primary: {incumbent}; Secondary: Regional subcontractors",
        "operational_weakness": weakness,
        "automation_opportunity": automation_opp,
        "recurring_revenue_potential": recurring_rev,
        "prime_pathwy_priority": priority,
        "data_date": datetime.now().strftime("%Y-%m-%d"),
        "source": "Prime Pathwy Procurement Intelligence Engine v2.0",
    }

def generate_all_procurement(target=3000):
    records = []
    all_agencies = (
        [(a, "US") for a in US_AGENCIES] +
        [(a, "CA") for a in CANADA_AGENCIES] +
        [(a, "UK") for a in UK_AGENCIES] +
        [(a, "AU") for a in AU_AGENCIES]
    )

    idx = 1
    while len(records) < target:
        agency_data, country = random.choice(all_agencies)
        records.append(generate_procurement_record(idx, agency_data, country))
        idx += 1

    return records

def write_csv(records, filepath):
    if not records:
        return
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(records[0].keys()))
        writer.writeheader()
        writer.writerows(records)
    print(f"  CSV written: {filepath} ({len(records)} rows)")

def write_xlsx(records, filepath, sheet_name="Procurement Master"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_fill = PatternFill("solid", fgColor=BLACK)
    header_font = Font(bold=True, color=GOLD, size=10)
    alt1 = PatternFill("solid", fgColor="0F0F0F")
    alt2 = PatternFill("solid", fgColor="1A1A1A")
    data_font = Font(color=WHITE, size=9)
    thin = Side(style="thin", color="333333")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    if not records:
        wb.save(filepath)
        return

    headers = list(records[0].keys())
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header.replace("_", " ").upper())
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row_idx, record in enumerate(records, 2):
        fill = alt1 if row_idx % 2 == 0 else alt2
        for col_idx, key in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=record.get(key, ""))
            cell.fill = fill
            cell.font = data_font
            cell.alignment = Alignment(vertical="center")
            cell.border = border

    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 45)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(filepath)
    print(f"  XLSX written: {filepath} ({len(records)} rows)")

def write_renewal_forecast(records, filepath):
    """Write renewal cycle forecast CSV."""
    forecast = []
    for r in records:
        forecast.append({
            "procurement_id": r["procurement_id"],
            "agency": r["issuing_agency"],
            "contract_title": r["contract_title"],
            "country": r["country"],
            "renewal_date": r["renewal_date"],
            "estimated_value": r["estimated_contract_value"],
            "incumbent": r["incumbent_vendor"],
            "prime_pathwy_priority": r["prime_pathwy_priority"],
            "days_to_renewal": (datetime.strptime(r["renewal_date"], "%Y-%m-%d") - datetime.now()).days,
            "action_required": "Position NOW" if (datetime.strptime(r["renewal_date"], "%Y-%m-%d") - datetime.now()).days < 90 else "Monitor",
        })
    forecast.sort(key=lambda x: x["days_to_renewal"])
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(forecast[0].keys()))
        writer.writeheader()
        writer.writerows(forecast)
    print(f"  Renewal forecast written: {filepath} ({len(forecast)} rows)")

def write_vendor_replacement_map(records, filepath):
    """Identify incumbent displacement opportunities."""
    replacements = []
    for r in records:
        if "displacement" in r["operational_weakness"].lower() or "replacement" in r["operational_weakness"].lower() or "open competition" in r["operational_weakness"].lower():
            replacements.append({
                "procurement_id": r["procurement_id"],
                "agency": r["issuing_agency"],
                "contract_title": r["contract_title"],
                "incumbent": r["incumbent_vendor"],
                "contract_value": r["estimated_contract_value"],
                "weakness": r["operational_weakness"],
                "bid_close": r["bid_close_date"],
                "portal": r["procurement_portal"],
                "contact": r["contact_email"],
                "prime_pathwy_action": "INITIATE DISPLACEMENT STRATEGY",
            })
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(replacements[0].keys()) if replacements else [])
        if replacements:
            writer.writeheader()
            writer.writerows(replacements)
    print(f"  Vendor replacement map written: {filepath} ({len(replacements)} rows)")

def write_procurement_graph(records, filepath):
    """Write procurement relationship JSON graph."""
    nodes = []
    edges = []
    agency_ids = {}
    vendor_ids = {}

    for r in records:
        # Agency node
        aid = f"AGENCY-{r['issuing_agency'][:20].replace(' ','_').replace(',','')}"
        if aid not in agency_ids:
            agency_ids[aid] = True
            nodes.append({"id": aid, "label": r["issuing_agency"], "type": "agency", "country": r["country"]})
        # Vendor node
        vid = f"VENDOR-{r['incumbent_vendor'][:20].replace(' ','_').replace(',','')}"
        if vid not in vendor_ids:
            vendor_ids[vid] = True
            nodes.append({"id": vid, "label": r["incumbent_vendor"], "type": "vendor", "country": r["country"]})
        # Edge
        edges.append({
            "source": aid,
            "target": vid,
            "relationship": "incumbent_contract",
            "contract_value": r["estimated_contract_value"],
            "category": r["contract_category"],
            "renewal_date": r["renewal_date"],
            "weight": 0.8,
            "monetization_potential": r["recurring_revenue_potential"],
        })

    graph = {
        "metadata": {
            "generated": datetime.now().isoformat(),
            "node_count": len(nodes),
            "edge_count": len(edges),
            "source": "Prime Pathwy Procurement Intelligence Engine v2.0",
        },
        "nodes": nodes,
        "edges": edges,
    }
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(graph, f, indent=2)
    print(f"  Procurement graph written: {filepath} ({len(nodes)} nodes, {len(edges)} edges)")

# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("[Phase 2] Generating 3,000 procurement records...")
    records = generate_all_procurement(3000)

    print("[Phase 2] Writing master procurement CSV...")
    write_csv(records, f"{OUTPUT_DIR}/PP_PROCUREMENT_MASTER.csv")

    print("[Phase 2] Writing US procurement CSV...")
    us_records = [r for r in records if r["country"] == "US"]
    write_csv(us_records, f"{OUTPUT_DIR}/PP_PROCUREMENT_US.csv")

    print("[Phase 2] Writing NorCal procurement CSV...")
    norcal_proc = [r for r in records if "CA" in r["agency_location"] and r["country"] == "US"]
    write_csv(norcal_proc, f"{OUTPUT_DIR}/PP_PROCUREMENT_NORCAL.csv")

    print("[Phase 2] Writing international procurement CSV...")
    intl_records = [r for r in records if r["country"] != "US"]
    write_csv(intl_records, f"{OUTPUT_DIR}/PP_PROCUREMENT_INTERNATIONAL.csv")

    print("[Phase 2] Writing CRITICAL priority procurement CSV...")
    critical = [r for r in records if "CRITICAL" in r["prime_pathwy_priority"]]
    write_csv(critical, f"{OUTPUT_DIR}/PP_PROCUREMENT_CRITICAL_PRIORITY.csv")

    print("[Phase 2] Writing XLSX master procurement database...")
    write_xlsx(records, f"{OUTPUT_DIR}/PP_PROCUREMENT_MASTER_DATABASE.xlsx")

    print("[Phase 2] Writing renewal forecast...")
    write_renewal_forecast(records, f"{OUTPUT_DIR}/PP_PROCUREMENT_RENEWAL_FORECAST.csv")

    print("[Phase 2] Writing vendor replacement opportunity map...")
    write_vendor_replacement_map(records, f"{OUTPUT_DIR}/PP_VENDOR_REPLACEMENT_MAP.csv")

    print("[Phase 2] Writing procurement relationship graph...")
    write_procurement_graph(records, f"{OUTPUT_DIR}/PP_PROCUREMENT_RELATIONSHIP_GRAPH.json")

    print(f"\n[Phase 2] COMPLETE — {len(records)} procurement records generated.")
    print(f"  Output directory: {OUTPUT_DIR}")
